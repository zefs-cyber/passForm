# myapp/views.py
from django.shortcuts import render, redirect
from .forms import UserFill
from .models import FormData
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from datetime import datetime
from django.core.files.storage import default_storage
from django.contrib.auth.decorators import login_required
from django.conf import settings
from django.contrib.auth import authenticate, login
from django.http import HttpResponse
from django.http import FileResponse
from .utils import zip_media_folder
import pandas as pd
from io import BytesIO
import os
import zipfile
import shutil


def index(request, language='tj'):
    if request.method == 'POST':
        form = UserFill(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            print(f"{form.cleaned_data['fullNameCyrillic']} - accepted")  # Corrected quotes
            print(form.cleaned_data)
            save_to_excel(form.cleaned_data)
            template_name = f"s_{language}.html"  
            return render(request, template_name)
    else:
        form = UserFill()
    if language in ['ru', 'tj', 'en']:
        template_name = f"{language}.html"  
        print(template_name)
        return render(request, template_name, {'form': form})
    return True

def replace_tajik_characters(string):
    replacement_dict = {'Ғ': 'Г', 'ғ': 'г', 'Қ': 'К', 'қ': 'к', 'Ҷ': 'Ч', 'ҷ': 'ч',
                        'Ҳ': 'Х', 'ҳ': 'х', 'Ӣ': 'и', 'ӣ': 'и', 'Ӯ': 'У', 'ӯ': 'у'}
    modified_string = ""
    if string != None:
        for char in string:
            if char in replacement_dict:
                modified_string += replacement_dict[char]
            else:
                modified_string += char

    return modified_string

def append_image(image_path, new_width=None, new_height=None):
    existing_workbook = load_workbook('myapp\\media\\student_data.xlsx')
    existing_worksheet = existing_workbook.active
    img = Image(image_path)

    if new_width is not None:
        img.width = new_width
    if new_height is not None:
        img.height = new_height

    cell = f'G{existing_worksheet.max_row}' 

    # Extracting the row number from the cell string
    row_number = int(cell[1])

    # Set the height of the row using the row number
    existing_worksheet.row_dimensions.height = new_height

    existing_worksheet.add_image(img, cell)
    existing_workbook.save('myapp\\media\\student_data.xlsx')

def change_cell_size(file_path, height):
    # Load the workbook
    workbook = load_workbook(file_path)

    # Access the first (and only) sheet
    sheet = workbook.active

    # Set the height for all rows
    for row_index, row in enumerate(sheet.iter_rows(), start=1):
        for cell in row:
            if row_index == 1:
                sheet.row_dimensions[cell.row].height = 20
            else:
                sheet.row_dimensions[cell.row].height = height

    # Set the width to autofit for all columns
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    # Save the changes
    workbook.save(file_path)

def delete_all_rows_except_header():

    file_paths = [
        'myapp\\media\\form_data.xlsx',
        'myapp\\media\\student_data.xlsx'
    ]

    for file_path in file_paths:
        try:
            # Load the Excel file into a DataFrame
            df = pd.read_excel(file_path, header=0)

            # Keep only the header row
            df = pd.DataFrame(columns=df.columns)

            # Save the modified DataFrame back to the Excel file
            df.to_excel(file_path, index=False)

        except Exception as e:
            print(f"Error processing {file_path}: {str(e)}")

def delete_contents_of_subfolders(folder_path):
    # Iterate through all subfolders in the given folder
    for root, dirs, files in os.walk(folder_path):
        for folder in dirs:
            folder_path = os.path.join(root, folder)
            
            # Delete all files inside the subfolder
            for filename in os.listdir(folder_path):
                file_path = os.path.join(folder_path, filename)
                try:
                    if os.path.isfile(file_path):
                        os.unlink(file_path)
                    elif os.path.isdir(file_path):
                        shutil.rmtree(file_path)
                except Exception as e:
                    print(f"Failed to delete {file_path}. Reason: {e}")

def download_media(request):
    # Define the base directory for media files
    media_base_dir = 'myapp/media'
    change_cell_size('myapp/media/student_data.xlsx', 100)

    # Create a BytesIO buffer to store the zip file
    zip_buffer = BytesIO()

    # Create a ZipFile object
    with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
        # Iterate through all files in the media directory and add them to the zip file
        for root, dirs, files in os.walk(media_base_dir):
            for file in files:
                file_path = os.path.join(root, file)
                # Specify the arcname to remove the initial path from the zip file
                zip_file.write(file_path, arcname=os.path.relpath(file_path, media_base_dir))

    # Seek to the beginning of the buffer
    zip_buffer.seek(0)

    # Create a FileResponse and return it for download
    response = FileResponse(zip_buffer, content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename=media_files.zip'
    # delete_all_rows_except_header()
    # delete_contents_of_subfolders(media_base_dir)
    print('All Data has been downloaded')
    # queryset = FormData.objects.all()
    # queryset.delete()
    return response

def save_to_excel(form_data):
    try:
        # Try to load the existing workbook
        wb_form = load_workbook('myapp\\media\\form_data.xlsx')
        wb_student = load_workbook('myapp\\media\\student_data.xlsx')
    except FileNotFoundError:
        # If the file doesn't exist, create a new workbook
        wb_form = Workbook()
        wb_student = Workbook()

    ws_form = wb_form.active
    ws_student = wb_student.active

    # Header row (only if the file is newly created)
    if not ws_form.rows:
        header_form = [
                "#",                                            #0 Index
                "ФИО",                                          #1 ФИО
                "Пол",                                          #2 Пол
                "ФИО на латинице",                              #3 ФИО на латинице
                "Дата рождения",                                #4 Дата рождения
                "Гражданство Таджикистана (1 - Да, 0 - Нет)",   #5 Гражданство Таджикистана (1 - Да, 0 - Нет)
                "Если RESIDENT 0 - COUNTRY обязательное поле",  #6 Если RESIDENT 0 - COUNTRY обязательное поле
                "Серия паспорта",                               #7 Серия паспорта
                "Номер паспорта",                               #8 Номер паспорта
                "Дата выдачи",                                  #9 Дата выдачи
                "Дата истечения",                               #10 Дата истечения
                "Кем выдан паспорт",                            #11 Кем выдан паспорт
                "Номер телефона",                               #12 Номер телефона
                "ИНН",                                          #13 ИНН
                "userType",
                "Работа/Учеба",
                "Регион",
                "Секретный вопрос",                             #14 Секретный вопрос
                "Ответ",                                        #15 Ответ
                "Адрес проживания/прописки",                    #16 Адрес проживания/прописки
                "Имя на карте",                                 #17 Имя на карте
                "Заполняется сотрудником банка",                #20 Заполняется сотрудником банка
                "Заполняется сотрудником банка",                #21 Заполняется сотрудником банка
                "",                                             #22 ????????
                "Заполняется сотрудником банка",                #23 Заполняется сотрудником банка
                "Зарплата сотрудников",                         #24 Зарплата сотрудников
            ]
        header_student = [
            "№", 
            "Ному насаб", 
            "Корти донишҷу", 
            "факультет/мактаб", 
            "Шакли таҳсил",
            "Нохия",
            "Фото (цветное 3х4)"]
        ws_form.append(header_form)
        ws_student.append(header_student)

    # Data row for wb_form
    card_name = form_data['fullNameLatin'].split(' ')[1] + " " + form_data['fullNameLatin'].split(' ')[0] 
    if len(card_name) > 19:
        card_name = f"{card_name.split(' ')[0][0]}. {card_name.split(' ')[1]}".upper()
    else:
        card_name = card_name.upper()
    
    birthday = datetime.strptime(str(form_data['birthdate']), "%Y-%m-%d")
    formatted_birthday = birthday.strftime("%d.%m.%Y")

    issuedate = datetime.strptime(str(form_data['issueDate']), "%Y-%m-%d")
    formatted_issuedate = issuedate.strftime("%d.%m.%Y")

    expirydate = datetime.strptime(str(form_data['expiryDate']), "%Y-%m-%d")
    formatted_expirydate = expirydate.strftime("%d.%m.%Y")

    if form_data['userType'] == 'Талаба':
        work_study = form_data['school']
    elif form_data['userType'] == 'Донишчу':
        work_study = form_data['university']
    else:
        work_study = form_data['company']

        
    
    data_row_form = [
        ws_form.max_row,
        replace_tajik_characters(form_data['fullNameCyrillic'].title()),
        form_data['gender'][0],
        form_data['fullNameLatin'].upper(),
        formatted_birthday,
        1,
        0,
        form_data['passportNumber'][0],
        form_data['passportNumber'][1:],
        formatted_issuedate,
        formatted_expirydate,
        replace_tajik_characters(form_data['jobLocation']),
        form_data['phoneNumber'],
        form_data['rma'],
        form_data['userType'],
        work_study,
        form_data['region'],
        'Имя',
        replace_tajik_characters(form_data['fullNameCyrillic'].split(' ')[1]),
        replace_tajik_characters(form_data['address']),
        card_name,
        "",
        "",
        "",
        "",
        "",]
    ws_form.append(data_row_form)

    if form_data['userType'] != 'Корманд':
    # Data row for wb_student
        data_row_student = [
            ws_student.max_row+1,
            replace_tajik_characters(form_data['fullNameCyrillic']),
            form_data['studentCard'],
            replace_tajik_characters(work_study),
            replace_tajik_characters(form_data['eduType']),
            form_data['region']
        ]
        ws_student.append(data_row_student)
        wb_student.save('myapp\\media\\student_data.xlsx')

        name = form_data['fullNameCyrillic'].replace(' ', '_')
        try:
            image_filename = f'{name}_photo3x4.jpg'
            image_path = os.path.join('myapp', 'media', '3x4', image_filename)
        except:
            try:
                image_filename = f'{name}_photo3x4.jpeg'
                image_path = os.path.join('myapp', 'media', '3x4', image_filename)
            except:
                image_filename = f'{name}_photo3x4.png'
                image_path = os.path.join('myapp', 'media', '3x4', image_filename)
            
        
        new_width = 75  
        new_height = 100
        append_image(image_path, new_width, new_height)

    # Save the files
    wb_form.save('myapp\\media\\form_data.xlsx')
    print(f"{form_data['fullNameCyrillic']} - saved")
  
def load_user_credentials(file_path):
    try:
        df = pd.read_excel(file_path)
        return df.set_index('username').to_dict('index')
    except FileNotFoundError:
        return {}
    except KeyError:
        return {}

def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        # Load user credentials from the Excel file

        file_path = 'myapp\\user_credentials.xlsx'
        user_credentials = load_user_credentials(file_path)
        
        # Authenticate user
        if username in user_credentials and user_credentials[username]['password'] == password:
            # Manually set the session for the user
            request.session['username'] = username
            return redirect('home')  # Redirect to home page after successful login
        else:
            return render(request, 'login.html', {'error': 'Invalid username or password.'})
    else:
        return render(request, 'login.html')

def home(request):
    # Ensure the user is authenticated
    if 'username' not in request.session:
        return redirect('login')

    username = request.session['username']
    
    # Load user credentials
    user_credentials_path = settings.BASE_DIR / 'myapp' / 'user_credentials.xlsx'
    user_credentials = load_user_credentials(user_credentials_path)

    if username not in user_credentials:
        return redirect('login')

    user_uni = user_credentials[username].get('uni')

    if not user_uni:
        return redirect('login')

    # Load the Excel file
    form_data_path = settings.BASE_DIR / 'myapp' / 'media' / 'form_data.xlsx'
    df = pd.read_excel(form_data_path)
    
    # Filter rows where 'Работа/Учеба' corresponds to the user's 'uni'
    df_filtered = df[df['Работа/Учеба'] == user_uni]
    
    # Combine 'Серия паспорта' and 'Номер паспорта' into one string
    df_filtered['passport_number'] = df_filtered['Серия паспорта'].astype(str) + df_filtered['Номер паспорта'].astype(str)
    
    # Extract the relevant columns
    form_data_parts = df_filtered[['ФИО', 'passport_number']].to_dict('records')
    
    # Print the extracted data for debugging
    print(f"table {form_data_parts}")

    # Render the template with the data
    return render(request, 'home.html', {'form_data_parts': form_data_parts})

def approve_rows(request):
    if request.method == 'POST':
        # Load the Excel file
        file_path = 'myapp\\media\\form_data.xlsx'
        df = pd.read_excel(file_path)
        
        # Combine 'Серия паспорта' and 'Номер паспорта' into one string
        df['passport_number'] = df['Серия паспорта'].astype(str) + df['Номер паспорта'].astype(str)

        # Get the selected rows from the form
        selected_rows = request.POST.getlist('selected_rows')
        
        # Filter the DataFrame to get only the selected rows
        approved_df = df[df['passport_number'].isin(selected_rows)]

        # Append the selected rows to approved_form_data.xlsx
        approved_file_path = 'myapp\\media\\approved_form_data.xlsx'
        
        try:
            approved_df_existing = pd.read_excel(approved_file_path)
            approved_df_combined = pd.concat([approved_df_existing, approved_df], ignore_index=True)
        except FileNotFoundError:
            approved_df_combined = approved_df

        approved_df_combined.to_excel(approved_file_path, index=False)

        # Remove approved rows from the original DataFrame
        df = df[~df['passport_number'].isin(selected_rows)]
        
        # Save the updated DataFrame back to the original file
        df.to_excel(file_path, index=False)

        # Reload the home view with a success message
        form_data_parts = df[['ФИО', 'passport_number']].to_dict('records')
        return render(request, 'home.html', {'form_data_parts': form_data_parts, 'message': 'Selected rows have been approved and saved.'})

    return redirect('home')  # Redirect back if not POST request
